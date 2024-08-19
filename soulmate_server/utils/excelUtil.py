# -*- coding: utf-8 -*-
import os

# 创建工作表并填充数据
from openpyxl import Workbook, load_workbook

from soulmate_server.common.mysql_tool import mysqlSession
from soulmate_server.models.energy import Order
from soulmate_server.models.user import User, LoginLog
from soulmate_server.utils.emailUtils import sendUserDataEmail


def create_sheet_and_fill_data(sheet_name, data_list, workbook):
    # 创建一个工作簿
    sheetCount = len(workbook.sheetnames)
    # 创建一个工作表
    sheet = workbook.create_sheet(title=sheet_name)

    # 写入字段名
    if data_list:
        field_names = list(data_list[0].keys())
        for col_num, field_name in enumerate(field_names, start=1):
            sheet.cell(row=1, column=col_num, value=field_name)

        # 写入数据
        for row_num, data in enumerate(data_list, start=2):
            for col_num, field_name in enumerate(field_names, start=1):
                sheet.cell(row=row_num, column=col_num, value=data.get(field_name, ''))


def downloadUserData(userId, email, sql: mysqlSession = None):
    user = sql.query(User).filter(User.userId == userId,User.status==0).first()
    json = {}
    json.update({"email": user.email, "name": user.nickName, "headImage": user.avatar, "energy": user.energy,
                 "emergencyContact": "Unopened" if user.emergencyContact == 0 else "Turned on",
                 "emergencyEmail": user.emergencyEmail})
    workbook = Workbook()
    create_sheet_and_fill_data("user", [json], workbook=workbook)
    orders = sql.query(Order).filter(Order.userId == userId,
                                     Order.status == 0).all()
    order_list = []
    for order in orders:
        order_list.append({
            "orderAmount": order.orderAmount,
            "productAmount": order.productAmount,
            "productEnergy": order.productEnergy,
            "productName": order.productName,
            "productType": order.productType,
            "type": "能量包" if order.type == 0 else "订阅" if order.type == 1 else "定制角色",
            "result": "进行中" if order.result == 0 else "成功" if order.result == 1 else "失败",
            "productNum": order.productNum,
            "productTypeName": "能量包" if order.productType == 0 else "订阅" if order.productType == 1 else "定制角色"
        })
    create_sheet_and_fill_data("order", order_list, workbook=workbook)
    logins = sql.query(LoginLog).filter(LoginLog.userId == userId).all()
    login_list = []
    for login in logins:
        login_list.append({
            "loginType": "email" if login.loginType == 0 else "google" if login.loginType == 1 else "facebook",
            "version": login.version,
            "platform": login.platform,
            "deviceModel": login.deviceModel,
            "status": "normal status" if login.loginType == 0 else "delete status"
        })
    create_sheet_and_fill_data("login", login_list, workbook=workbook)
    # 获取要删除的工作表
    sheet_to_remove = workbook.sheetnames[0]
    # 删除第一个工作表
    workbook.remove(workbook[sheet_to_remove])
    sendUserDataEmail(email=email, workbook=workbook, Subject="Your Soulmate ELF Account Data Is Ready for Download")







import openpyxl


def refresh_excel_functions(file_path):
    # 检查文件是否存在
    if not os.path.exists(file_path):
        print(f"文件 {file_path} 不存在。")
        return

    # 加载工作簿
    wb = load_workbook(file_path)
    # 刷新数据（使用openpyxl）
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for table in ws.tables.values():
            table.refresher()

    # 保存更改
    wb.save('D:/data/upload/client/excel/template/a123s.xlsx')
    print(f"文件 {file_path} 已刷新。")
#
# if __name__ == "__main__":
#     excel)
